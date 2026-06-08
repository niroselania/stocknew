from __future__ import annotations

import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_COLUMNS = [
    {"base": "CARRITO", "col": "C"},
    {"base": "LOCAL", "col": "L"},
    {"base": "MAYORISTA", "col": "M"},
    {"base": "RETAIL", "col": "R"},
    {"base": "BARI", "col": "B"},
    {"base": "DEPOSITO", "col": "D"},
    {"base": "EXPORTA", "col": "E"},
    {"base": "CONTROL", "col": "CO"},
    {"base": "PENDIENTES", "col": "P.PAGO"},
]


def normalize(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def as_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    cleaned = re.sub(r"[^\d.\-]", "", str(value or "").replace(",", "."))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def find_header_index(header: list[Any], name: str) -> int:
    wanted = normalize(name)
    for index, cell in enumerate(header):
        if normalize(cell) == wanted:
            return index
    return -1


def is_total_row(row: list[Any]) -> bool:
    return any(normalize(cell) == "TOTAL" for cell in row)


def sheet_rows(workbook_path: Path, sheet_name: str) -> list[list[Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        wanted = normalize(sheet_name)
        worksheet = None
        for name in workbook.sheetnames:
            if normalize(name) == wanted:
                worksheet = workbook[name]
                break
        if worksheet is None:
            raise ValueError(f"No encontré una pestaña llamada {sheet_name}.")
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def detect_pending_columns(pending_rows: list[list[Any]]) -> tuple[int, int]:
    for row in pending_rows[:10]:
        qty_index = -1
        for index, cell in enumerate(row):
            normalized = normalize(cell)
            if "QUANTITY" in normalized or "CANT" in normalized:
                qty_index = index
                break
        if qty_index >= 0:
            concat_index = 0
            for index, cell in enumerate(row):
                if "CONCAT" in normalize(cell):
                    concat_index = index
                    break
            return qty_index, concat_index
    return -1, 0


def build_stock_payload(workbook_path: Path) -> dict[str, Any]:
    stock_rows = sheet_rows(workbook_path, "STOCK")
    if not stock_rows:
        raise ValueError("La pestaña STOCK está vacía.")

    header = list(stock_rows[0])
    concat_col = find_header_index(header, "CONCAT")
    sku_col = find_header_index(header, "SKU")
    color_col = find_header_index(header, "COLOR")
    talle_col = find_header_index(header, "TALLE")
    nombre_col = find_header_index(header, "NOMBRE")
    linea_col = find_header_index(header, "LINEA")

    try:
        pending_rows = sheet_rows(workbook_path, "PENDIENTES")
    except ValueError:
        pending_rows = []

    pending_qty_col, pending_concat_col = detect_pending_columns(pending_rows)
    pending_by_concat: dict[str, float] = {}
    if pending_rows and pending_qty_col >= 0:
        for row in pending_rows[1:]:
            if not row:
                continue
            concat = normalize(row[pending_concat_col] if pending_concat_col < len(row) else "")
            if not concat:
                continue
            qty = as_number(row[pending_qty_col] if pending_qty_col < len(row) else 0)
            pending_by_concat[concat] = pending_by_concat.get(concat, 0.0) + qty

    index: dict[str, dict[str, Any]] = {}
    inventory: dict[str, dict[str, Any]] = {}
    data_rows: list[list[Any]] = []
    lineas: set[str] = set()

    for row in stock_rows[1:]:
        if is_total_row(row):
            continue

        row_values = list(row) + [""] * max(0, len(header) - len(row))
        concat_value = normalize(row_values[concat_col]) if concat_col >= 0 else ""
        built_value = normalize(
            " ".join(
                str(row_values[col])
                for col in (sku_col, color_col, talle_col)
                if col >= 0 and row_values[col] not in (None, "")
            )
        )
        key = concat_value or built_value
        if not key:
            continue

        bases: dict[str, float] = {}
        for base in BASE_COLUMNS:
            col_index = find_header_index(header, base["col"])
            qty = as_number(row_values[col_index]) if col_index >= 0 else 0.0
            bases[base["base"]] = qty

        pending_qty = pending_by_concat.get(key, 0.0)
        if pending_qty > 0:
            bases["PENDIENTES"] = pending_qty

        product = {
            "concat": key,
            "sku": row_values[sku_col] if sku_col >= 0 else "",
            "color": row_values[color_col] if color_col >= 0 else "",
            "talle": row_values[talle_col] if talle_col >= 0 else "",
            "nombre": row_values[nombre_col] if nombre_col >= 0 else "",
            "linea": row_values[linea_col] if linea_col >= 0 else "",
            "bases": bases,
            "row": ["" if value is None else value for value in row_values[: len(header)]],
        }

        linea = normalize(product["linea"])
        if linea:
            lineas.add(linea)

        data_rows.append(product["row"])
        index[key] = {
            "sku": product["sku"],
            "color": product["color"],
            "talle": product["talle"],
            "nombre": product["nombre"],
            "linea": product["linea"],
            "bases": bases,
        }
        inventory[key] = product

    totals: dict[str, float] = {base["base"]: 0.0 for base in BASE_COLUMNS}
    total_row = None
    for row in reversed(stock_rows[1:]):
        if is_total_row(row):
            total_row = list(row) + [""] * max(0, len(header) - len(row))
            break

    for base in BASE_COLUMNS:
        col_index = find_header_index(header, base["col"])
        if col_index < 0:
            continue
        if total_row is not None:
            totals[base["base"]] = as_number(total_row[col_index])
        else:
            totals[base["base"]] = sum(
                as_number(row[col_index]) if col_index < len(row) else 0.0 for row in data_rows
            )

    if pending_by_concat:
        totals["PENDIENTES"] = sum(pending_by_concat.values())

    return {
        "header": ["" if value is None else value for value in header],
        "rows": data_rows,
        "index": index,
        "inventory": inventory,
        "pendingByConcat": pending_by_concat,
        "totals": totals,
        "lineas": sorted(lineas),
        "productCount": len(data_rows),
        "pendingRows": [
            ["" if value is None else value for value in row]
            for row in pending_rows
        ],
        "pendingQuantityCol": pending_qty_col,
        "pendingConcatCol": pending_concat_col,
    }


def make_history_id(uploaded_at: int | None = None) -> str:
    timestamp = uploaded_at or int(datetime.now(tz=timezone.utc).timestamp())
    return datetime.fromtimestamp(timestamp, tz=timezone.utc).strftime("%Y%m%d_%H%M%S")


def compare_inventories(
    current: dict[str, dict[str, Any]],
    previous: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    changes: list[dict[str, Any]] = []
    keys = set(current) | set(previous)

    for key in sorted(keys):
        current_item = current.get(key)
        previous_item = previous.get(key)

        if current_item and not previous_item:
            for base, qty in current_item["bases"].items():
                if qty > 0:
                    changes.append(
                        {
                            "concat": key,
                            "sku": current_item.get("sku", ""),
                            "nombre": current_item.get("nombre", ""),
                            "base": base,
                            "before": 0.0,
                            "after": qty,
                            "delta": qty,
                            "type": "new",
                        }
                    )
            continue

        if previous_item and not current_item:
            for base, qty in previous_item["bases"].items():
                if qty > 0:
                    changes.append(
                        {
                            "concat": key,
                            "sku": previous_item.get("sku", ""),
                            "nombre": previous_item.get("nombre", ""),
                            "base": base,
                            "before": qty,
                            "after": 0.0,
                            "delta": -qty,
                            "type": "removed",
                        }
                    )
            continue

        current_bases = current_item["bases"]
        previous_bases = previous_item["bases"]
        for base in sorted(set(current_bases) | set(previous_bases)):
            before = float(previous_bases.get(base, 0.0))
            after = float(current_bases.get(base, 0.0))
            if before == after:
                continue
            changes.append(
                {
                    "concat": key,
                    "sku": current_item.get("sku", ""),
                    "nombre": current_item.get("nombre", ""),
                    "base": base,
                    "before": before,
                    "after": after,
                    "delta": after - before,
                    "type": "changed",
                }
            )

    summary = {
        "newProducts": sum(1 for item in changes if item["type"] == "new" and item["base"] == "LOCAL"),
        "removedProducts": sum(1 for item in changes if item["type"] == "removed" and item["base"] == "LOCAL"),
        "changedLines": len(changes),
        "positive": sum(1 for item in changes if item["delta"] > 0),
        "negative": sum(1 for item in changes if item["delta"] < 0),
    }

    return {"changes": changes, "summary": summary}
